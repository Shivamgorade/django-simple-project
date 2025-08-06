import os
# import pandas as pd
from django.http import JsonResponse
import openpyxl
from django.shortcuts import render, redirect
from django.conf import settings
from datetime import datetime

def form_view(request):
    if request.method == 'POST':
        # Metadata
        date = request.POST.get('date')  # ✅ renamed from 'date_time'
        line_no = request.POST.get('line_no')
        shift = request.POST.get('shift')
        part_no = request.POST.get('part_no')
        part_type = request.POST.get('part_type')
        supervisor = request.POST.get('supervisor')

        # ST-15 Inputs
        dmc_st15 = request.POST.get('st15_dmc_no')
        pos_501 = request.POST.get('POS_501')
        pos_502 = request.POST.get('POS_502')
        visual_staking_hu = request.POST.get('VISUAL_STAKING_HU')

        # ST-20 Inputs
        pos_311 = request.POST.get('pos311')
        pos_312 = request.POST.get('pos312')
        pos_411 = request.POST.get('pos411')
        pos_412 = request.POST.get('pos412')
        pos_531 = request.POST.get('pos531')
        pos_532 = request.POST.get('pos532')
        pos_541 = request.POST.get('pos541')
        pos_542 = request.POST.get('pos542')
        visual_staking_st20 = request.POST.get('visualStakingST20')

        # ST-25 Inputs
        pos_221 = request.POST.get('pos_221')
        visual_cbearing = request.POST.get('visual_cbearing')

        # ST-25 MV Staking
        pos201 = request.POST.get('pos201')
        pos202 = request.POST.get('pos202')
        pos211 = request.POST.get('pos211')
        pos212 = request.POST.get('pos212')
        visualStakingMV = request.POST.get('visualStakingMV')

        # ST-30 Motor Staking
        dmc_st30 = request.POST.get('dmc_st30')
        pos222 = request.POST.get('pos222')
        pos223 = request.POST.get('pos223')
        pos224 = request.POST.get('pos224')
        visual_staking_st30 = request.POST.get('visual_staking_st30')

        # ST-40 Inputs
        pos301 = request.POST.get('pos301')
        pos401 = request.POST.get('pos401')
        visualStakingST40 = request.POST.get('visualStakingST40')

        # ST-45 Inputs
        ecuScrewDamage = request.POST.get('ecuScrewDamage')
        pos231status = request.POST.get('pos231status')
        pos232status = request.POST.get('pos232status')

        # ST-85 Inputs
        ecuHousingDamage = request.POST.get('ecuHousingDamage')

        # LAST CHECKED BY Inputs
        checked_by = request.POST.get('checkedBy')

        # Excel path
        excel_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', 'checksheet_data.xlsx')

        # Create or load workbook
        if os.path.exists(excel_path):
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = [
                'DATE', 'LINE NO', 'SHIFT', 'PART NO', 'PART TYPE', 'SUPERVISOR',
                'ST15 DMC', 'POS 501 Depth', 'POS 502 Depth', 'Visual Staking HU (ST-15)',
                'POS 311 Depth', 'POS 312 Depth', 'POS 411 Depth', 'POS 412 Depth',
                'POS 531 Depth', 'POS 532 Depth', 'POS 541 Depth', 'POS 542 Depth',
                'Visual Staking (ST-20)',
                'ST25 DMC', 'Visual – C-Bearing',
                'MV DMC', '201 Depth', '202 Depth', '211 Depth', '212 Depth', 'MV Visual',
                'ST30 DMC', '222 Depth', '223 Depth', '224 Depth', 'Visual Staking (ST-30)',
                '301 Depth', '401 Depth', 'Visual Staking (ST-40)',
                'Pos. 231 Status', 'Pos. 232 Status', 'ECU Screw Head Damage',
                'ECU Housing Damage',
                'Checked By',
            ]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=3, column=col, value=header)

        # Row data
        row_data = [
            date, line_no, shift, part_no, part_type, supervisor,
            dmc_st15, pos_501, pos_502, visual_staking_hu,
            pos_311, pos_312, pos_411, pos_412,
            pos_531, pos_532, pos_541, pos_542,
            visual_staking_st20,
            pos_221, visual_cbearing,
            pos201, pos202, pos211, pos212, visualStakingMV,
            dmc_st30, pos222, pos223, pos224, visual_staking_st30,
            pos301, pos401, visualStakingST40,
            pos231status, pos232status, ecuScrewDamage,
            ecuHousingDamage,
            checked_by,
        ]

        # Find next available row
        next_row = 4
        while sheet.cell(row=next_row, column=1).value:
            next_row += 1

        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=next_row, column=col, value=value)

        workbook.save(excel_path)
        return redirect('form_view')

    return render(request, 'form.html')


def dashboard_view(request):
    return render(request, 'dashboard.html')


from django.http import JsonResponse
import os
import traceback
import pandas as pd
from django.conf import settings


def load_and_filter_excel(date, line_no, shift):
    """Load the Excel file and filter by date, line number, and shift"""
    try:
        path = os.path.join(settings.BASE_DIR, 'myapp', 'data', 'checksheet_data.xlsx')
        df = pd.read_excel(path, header=2)
        df.columns = [str(col).strip() for col in df.columns]

        df['DATE'] = pd.to_datetime(df['DATE']).dt.date.astype(str)
        df['LINE NO'] = df['LINE NO'].astype(str).str.strip().str.upper()
        df['SHIFT'] = df['SHIFT'].astype(str).str.strip().str.lower()

        filtered_df = df[
            (df['DATE'] == date) &
            (df['LINE NO'] == line_no) &
            (df['SHIFT'] == shift)
        ]

        return filtered_df
    except Exception as e:
        raise RuntimeError(f"Error loading or filtering Excel: {str(e)}")


def extract_positions(row, positions):
    """Extract values for given positions from a row"""
    labels, values = [], []
    for pos in positions:
        if pos in row and pd.notna(row[pos]):
            labels.append(pos)
            values.append(round(float(row[pos]), 2))
    return labels, values


# Chart ST-15
def get_filtered_chart_data(request):
    try:
        date = request.GET.get('date', '').strip()
        line_no = request.GET.get('line', '').strip().upper()
        shift = request.GET.get('shift', '').strip().lower()

        if not date or not line_no or not shift:
            return JsonResponse({'message': 'Missing one or more required parameters'}, status=400)

        filtered_df = load_and_filter_excel(date, line_no, shift)
        if filtered_df.empty:
            return JsonResponse({'message': 'No data found for given filter!'}, status=404)

        row = filtered_df.iloc[0]
        labels, data = extract_positions(row, ['501', '502'])

        if not labels:
            return JsonResponse({'message': 'No valid depth data for POS 501 or 502'}, status=204)

        return JsonResponse({'labels': labels, 'data': data})

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'message': f'Internal Server Error: {str(e)}'}, status=500)


# Chart ST-20

def get_filtered_chart_data_st20(request):
    try:
        date = request.GET.get('date', '').strip()
        line_no = request.GET.get('line', '').strip().upper()
        shift = request.GET.get('shift', '').strip().lower()

        if not date or not line_no or not shift:
            return JsonResponse({'message': 'Missing one or more required parameters'}, status=400)

        filtered_df = load_and_filter_excel(date, line_no, shift)
        if filtered_df.empty:
            return JsonResponse({'message': 'No data found for given filter!'}, status=404)

        row = filtered_df.iloc[0]
        labels, values = extract_positions(row, ['311', '312', '411', '412', '531', '532', '541', '542'])

        if not labels:
            return JsonResponse({'message': 'No valid depth values for ST-20 positions'}, status=204)

        return JsonResponse({'labels': labels, 'values': values})

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'message': f'Internal Server Error: {str(e)}'}, status=500)


# Chart ST-25-C (only 1 position)
def get_filtered_chart_data_st25c(request):
    try:
        date = request.GET.get('date', '').strip()
        line_no = request.GET.get('line', '').strip().upper()
        shift = request.GET.get('shift', '').strip().lower()

        if not date or not line_no or not shift:
            return JsonResponse({'message': 'Missing one or more required parameters'}, status=400)

        filtered_df = load_and_filter_excel(date, line_no, shift)
        if filtered_df.empty:
            return JsonResponse({'message': 'No data found for given filter'}, status=404)

        row = filtered_df.iloc[0]
        value = row.get('221', None)

        if value is None or pd.isna(value):
            return JsonResponse({'message': 'No depth value found for position 221'}, status=204)

        return JsonResponse({'labels': ['221'], 'values': [round(float(value), 2)]})

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'message': f'Internal Server Error: {str(e)}'}, status=500)
    
def get_filtered_chart_data_st25mv(request):
    try:
        # --- Get and validate filters ---
        date = request.GET.get('date')
        line_no = request.GET.get('line')
        shift = request.GET.get('shift')

        if not date or not line_no or not shift:
            return JsonResponse({'message': 'Missing query parameters: date, line, shift'}, status=400)

        # --- Normalize filters ---
        date = date.strip()
        line_no = line_no.strip().upper()
        shift = shift.strip().lower()

        # --- Load Excel and prepare ---
        excel_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', 'checksheet_data.xlsx')
        df = pd.read_excel(excel_path, header=2)
        df.columns = [str(col).strip() for col in df.columns]

        df['DATE'] = pd.to_datetime(df['DATE']).dt.date.astype(str)
        df['LINE NO'] = df['LINE NO'].astype(str).str.strip().str.upper()
        df['SHIFT'] = df['SHIFT'].astype(str).str.strip().str.lower()

        # --- Filter data ---
        filtered_df = df[
            (df['DATE'] == date) &
            (df['LINE NO'] == line_no) &
            (df['SHIFT'] == shift)
        ]

        if filtered_df.empty:
            return JsonResponse({'message': 'No data found for given filter'}, status=404)

        row = filtered_df.iloc[0]

        # --- Positions ---
        positions = ['201', '202', '211', '212']
        labels = []
        values = []

        for pos in positions:
            if pos in row and pd.notna(row[pos]):
                labels.append(pos)
                values.append(round(float(row[pos]), 2))

        if not labels:
            return JsonResponse({'message': 'No valid ST-25-MV data found'}, status=204)

        return JsonResponse({'labels': labels, 'values': values})

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'message': f'Internal Server Error: {str(e)}'}, status=500)


def get_filtered_chart_data_st30(request):
    try:
        # Extract filters from query params
        date = request.GET.get('date')
        line_no = request.GET.get('line')
        shift = request.GET.get('shift')

        # Validate inputs
        if not date or not line_no or not shift:
            return JsonResponse({'message': 'Missing query parameters: date, line, shift'}, status=400)

        # Normalize inputs
        date = date.strip()
        line_no = line_no.strip().upper()
        shift = shift.strip().lower()

        # Load Excel
        excel_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', 'checksheet_data.xlsx')
        df = pd.read_excel(excel_path, header=2)
        df.columns = [str(col).strip() for col in df.columns]

        # Normalize filter columns
        df['DATE'] = pd.to_datetime(df['DATE']).dt.date.astype(str)
        df['LINE NO'] = df['LINE NO'].astype(str).str.strip().str.upper()
        df['SHIFT'] = df['SHIFT'].astype(str).str.strip().str.lower()

        # Filter data
        filtered_df = df[
            (df['DATE'] == date) &
            (df['LINE NO'] == line_no) &
            (df['SHIFT'] == shift)
        ]

        if filtered_df.empty:
            return JsonResponse({'message': 'No data found for given filter'}, status=404)

        row = filtered_df.iloc[0]

        # Positions for ST-30
        positions = ['222', '223', '224']
        labels = []
        values = []

        for pos in positions:
            if pos in row and pd.notna(row[pos]):
                labels.append(pos)
                values.append(round(float(row[pos]), 2))

        if not labels:
            return JsonResponse({'message': 'No valid ST-30 data found'}, status=204)

        return JsonResponse({'labels': labels, 'values': values})

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'message': f'Internal Server Error: {str(e)}'}, status=500)
    

def get_filtered_chart_data_st40(request):
    try:
        # Extract filters
        date = request.GET.get('date')
        line_no = request.GET.get('line')
        shift = request.GET.get('shift')

        # Validate
        if not date or not line_no or not shift:
            return JsonResponse({'message': 'Missing query parameters: date, line, shift'}, status=400)

        # Normalize
        date = date.strip()
        line_no = line_no.strip().upper()
        shift = shift.strip().lower()

        # Load Excel
        excel_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', 'checksheet_data.xlsx')
        df = pd.read_excel(excel_path, header=2)
        df.columns = [str(col).strip() for col in df.columns]

        df['DATE'] = pd.to_datetime(df['DATE']).dt.date.astype(str)
        df['LINE NO'] = df['LINE NO'].astype(str).str.strip().str.upper()
        df['SHIFT'] = df['SHIFT'].astype(str).str.strip().str.lower()

        # Filter
        filtered_df = df[
            (df['DATE'] == date) &
            (df['LINE NO'] == line_no) &
            (df['SHIFT'] == shift)
        ]

        if filtered_df.empty:
            return JsonResponse({'message': 'No data found for given filter'}, status=404)

        row = filtered_df.iloc[0]

        # ST-40 positions
        positions = ['301', '401']
        labels = []
        values = []

        for pos in positions:
            if pos in row and pd.notna(row[pos]):
                labels.append(pos)
                values.append(round(float(row[pos]), 2))

        if not labels:
            return JsonResponse({'message': 'No valid ST-40 depth data found'}, status=204)

        return JsonResponse({'labels': labels, 'values': values})

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'message': f'Internal Server Error: {str(e)}'}, status=500)


from django.http import JsonResponse
import pandas as pd
import os
from django.conf import settings

def info_card_api(request):
    try:
        excel_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', 'checksheet_data.xlsx')
        df = pd.read_excel(excel_path, header=2)

        df.columns = [str(c).strip() for c in df.columns]

        date = request.GET.get('date', '').strip()
        line = request.GET.get('line', '').strip().upper()
        shift = request.GET.get('shift', '').strip().lower()

        df['DATE'] = pd.to_datetime(df['DATE']).dt.date.astype(str)
        df['LINE NO'] = df['LINE NO'].astype(str).str.strip().str.upper()
        df['SHIFT'] = df['SHIFT'].astype(str).str.strip().str.lower()

        filtered = df[
            (df['DATE'] == date) &
            (df['LINE NO'] == line) &
            (df['SHIFT'] == shift)
        ]

        if filtered.empty:
            return JsonResponse({"message": "No data found"}, status=404)

        row = filtered.iloc[0]

        # ✅ Convert all values to str() to avoid TypeError
        data = {
            "DATE": str(row.get('DATE', '')),
            "LINE NO": str(row.get('LINE NO', '')),
            "SHIFT": str(row.get('SHIFT', '')),
            "PART NO": str(row.get('PART NO', '')),
            "PART TYPE": str(row.get('PART TYPE', '')),
            "SUPERVISOR": str(row.get('SUPERVISOR', '')),
            "ST-15 DMC NO": str(row.get('ST-15 DMC NO', '')),
            "ST-30 DMC NO": str(row.get('ST-30 DMC NO', '')),
            "CHECKED BY": str(row.get('CHECKED BY', '')),
        }

        return JsonResponse(data)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
    
def total_inspections_api(request):
    try:
        excel_path = os.path.join(settings.BASE_DIR, 'myapp', 'data', 'checksheet_data.xlsx')
        df = pd.read_excel(excel_path, header=2)  # Skips top 2 rows, third becomes header

        total_rows = df.shape[0]  # Total data rows (after header row)
        return JsonResponse({'total': total_rows})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
    




